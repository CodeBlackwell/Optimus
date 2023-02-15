import traceback
import copy

from compare_reports import Comparison
import sources
import warnings
import asyncio
import inspect
import arrow
import ast
import sys
import time
import pandas as pd
from pprint import pprint
from string import ascii_lowercase
import gspread
import gspread_formatting
from oauth2client.service_account import ServiceAccountCredentials
from openpyxl import load_workbook
import datetime
import xlsxwriter
from gspread_formatting import *

######### Globals

warnings.filterwarnings("ignore")
true = True
false = False
semaphore = 1
days_summary_dict = {"interval": "day", "comparisons": []}
weeks_summary_dict = {"interval": "week", "comparisons": []}
months_summary_dict = {"interval": "month", "comparisons": []}
total_matches = 0
total_edw3_mismatches = 0
total_edw2_mismatches = 0
total_edw2_orphans = 0
total_edw3_orphans = 0

g_utils = gspread.utils
scope = ["https://spreadsheets.google.com/feeds", 'https://www.googleapis.com/auth/spreadsheets',
         "https://www.googleapis.com/auth/drive.file", "https://www.googleapis.com/auth/drive"]
creds = ServiceAccountCredentials.from_json_keyfile_name("../creds.json", scope)
client = gspread.authorize(creds)
summary_sheet_names = [
    'month_summary',
    'week_summary',
    'day_summary'
]
inverted_summary_sheet_names = [
    'month_summary_inverted',
    'week_summary_inverted',
    'day_summary_inverted'
]

weeks_comp = []
days_comp = []
months_comp = []
myinterval = 0

arrow_formats = {
    "mm_dd_yyyy": "%m/%d/%Y",
    "yyyymmdd": "%Y%m%d"
}


##################

def generate_dates(starting_date, ending_date, interval='day'):
    result = {"dates": [],
              "interval": interval}
    for r in arrow.Arrow.range(interval, starting_date, ending_date):
        result["dates"].append(r.format('MM/DD/YYYY'))
    # check if the last week covers to the last day of the month
    last_date = result["dates"][-1]
    start = arrow.get(starting_date)
    end = arrow.get(ending_date)
    date_split = last_date.split("/")
    # If the last day is not == to the last day of the range,
    missing_days = (arrow.get(int(date_split[2]), int(date_split[0]), int(date_split[1])).is_between(start, end))
    if missing_days:
        # add the remaining days for the range
        result["dates"].append(ending_date.strftime('%m/%d/%Y'))
    return result


def check_replace(request_object, dates_hash, interval, date_idx, date_format=None):
    # extrapolate format info
    if date_format is None:
        date_format = {
            "filter_format": "mm_dd_yyyy",
            "dimension_format": "yyyymmdd",
            "date_col_name": "DATE"
        }
    column_name = date_format["date_col_name"]
    filter_field = "dim_date-%s" % date_format["filter_format"]
    dimension_field = "dim_date-%s" % date_format["dimension_format"]
    filter_format = arrow_formats[date_format["filter_format"]]

    # create a fresh filter
    start = dates_hash["dates"][date_idx]
    valid_start = start.split('/')
    valid_start = arrow.get(int(valid_start[2]), int(valid_start[0]), int(valid_start[1]))
    new_start = valid_start.strftime(filter_format)

    if interval != 'day':
        end = dates_hash["dates"][date_idx + 1]
        valid_end = end.split('/')
        valid_end = arrow.get(int(valid_end[2]), int(valid_end[0]), int(valid_end[1]))
        new_end = valid_end.strftime(filter_format)

    if interval == 'day':
        replacement_filter = {
            "field": filter_field,
            "op": "eq",
            "values": [
                new_start
            ]
        }
    elif interval == 'week':
        # 1st iteration specification
        if date_idx == 0:
            replacement_filter = {
                "field": filter_field,
                "op": "between",
                "values": [
                    new_start,
                    new_end
                ]
            }
        else:
            split_start = start.split('/')
            shiftable_date_obj = arrow.get(int(split_start[2]), int(split_start[0]), int(split_start[1]))
            modified_start = shiftable_date_obj.shift(days=+1).strftime(filter_format)
            replacement_filter = {
                "field": filter_field,
                "op": "between",
                "values": [
                    modified_start,
                    new_end
                ]
            }
    elif interval == 'month':
        split_end = end.split('/')
        shiftable_date_obj = arrow.get(int(split_end[2]), int(split_end[0]), int(split_end[1]))
        modified_end = shiftable_date_obj.shift(days=-1).strftime(filter_format)
        # last iteration specification
        if date_idx == len(dates_hash["dates"]) - 2:
            replacement_filter = {
                "field": filter_field,
                "op": "between",
                "values": [
                    new_start,
                    new_end
                ]
            }

        else:
            replacement_filter = {
                "field": filter_field,
                "op": "between",
                "values": [
                    new_start,
                    modified_end
                ]
            }

    replacement_date_dim = {
        "id": dimension_field,
        "name": column_name,
        "alias": "date"
    }
    # Look in the filters
    dim_name = 'dim_date'
    for report_name in request_object:
        for col_idx, col_obj in enumerate(request_object[report_name]["cols"]):
            if 'prepared_id' in col_obj:
                continue
            if dim_name in col_obj["id"]:
                # print("Deleting - Dimension - ", col_obj)
                # print("Replacement - Dimension - ", replacement_date_dim)
                del request_object[report_name]["cols"][col_idx]
        request_object[report_name]["cols"].append(replacement_date_dim)
        # Replace Filter
        for filter_idx, filter_obj in enumerate(request_object[report_name]["filters"]):
            if dim_name in filter_obj["field"]:
                # print("Deleting - Filter - ", filter_obj)
                # print("Replacement - Filter - ", replacement_filter)
                del request_object[report_name]["filters"][filter_idx]
        # Replace the Deleted (if found) filter
        request_object[report_name]["filters"].append(replacement_filter)
    # print("new object --- ", json.dumps(request_object))


def simple_comparison(edw2_ro, edw3_ro):
    edw2_url = 'https://picker-dev.avantlink.com/rpt'
    edw3_url = 'https://picker-shard.avantlink.com/rpt'
    Comparison(sources.PickerReport(picker_url=edw3_url,
                                    report_name="edw3_report_name",
                                    request_object=edw3_ro),
               sources.PickerReport(picker_url=edw2_url,
                                    report_name="edw2_report_name",
                                    request_object=edw2_ro)
               )


async def run_comparison(dates_hash, date_idx, edw2_request_object, edw3_request_object,
                         side_by_side=None,
                         interval="day", report_name="custom_comparison",
                         output_xlsx=False, cascade=False, cascade_start_date=None, cascade_end_date=None,
                         picker_address=None):
    if date_idx == len(dates_hash["dates"]) - 1 and interval != 'day':
        print('hitting end condition \n \n \n')
        return
    start = dates_hash["dates"][date_idx]
    split_start = start.split('/')
    shiftable_start_date_obj = arrow.get(int(split_start[2]), int(split_start[0]), int(split_start[1]))
    end = dates_hash["dates"][date_idx]
    try:
        end = dates_hash["dates"][date_idx + 1]
    except IndexError:
        pass
    split_end = end.split('/')
    shiftable_end_date_obj = arrow.get(int(split_end[2]), int(split_end[0]), int(split_end[1]))

    async with sem:
        if side_by_side:
            check_replace(edw2_request_object, dates_hash, interval, date_idx, side_by_side["date_format"])
            check_replace(edw3_request_object, dates_hash, interval, date_idx, side_by_side["date_format"])
        else:
            check_replace(edw2_request_object, dates_hash, interval, date_idx)
            check_replace(edw3_request_object, dates_hash, interval, date_idx)
        if interval == 'day':
            edw2_report_name = report_name + "__" + start + "_edw2"
            edw3_report_name = report_name + "__" + start + "_edw3"
        elif interval == 'week':
            if date_idx == 0:
                edw2_report_name = report_name + "__" + start + "--" + end + "__" + interval + "_edw2"
                edw3_report_name = report_name + "__" + start + "--" + end + "__" + interval + "_edw3"
            else:
                modified_start = shiftable_start_date_obj.shift(days=+1).strftime('%m/%d/%Y')
                edw2_report_name = report_name + "__" + modified_start + "--" + end + "__" + interval + "_edw2" + "--"
                edw3_report_name = report_name + "__" + modified_start + "--" + end + "__" + interval + "_edw3"
        elif interval == 'month':
            if date_idx == len(dates_hash["dates"]) - 2:
                edw2_report_name = report_name + "__" + start + "--" + end + "__" + interval + "_edw2" + "--"
                edw3_report_name = report_name + "__" + start + "--" + end + "__" + interval + "_edw3"
            else:
                modified_end = shiftable_end_date_obj.shift(days=-1).strftime('%m/%d/%Y')
                edw2_report_name = report_name + "__" + start + "--" + modified_end + "__" + interval + "_edw2" + "--"
                edw3_report_name = report_name + "__" + start + "--" + modified_end + "__" + interval + "_edw3"

        edw2_url = 'https://picker-dev.avantlink.com/rpt'
        edw3_url = 'https://picker-shard.avantlink.com/rpt'
        report_a_address = edw3_url
        report_b_address = edw2_url
        if picker_address == "edw3":
            report_a_address = edw3_url
            report_b_address = edw3_url
        if picker_address == "edw2":
            report_a_address = edw2_url
            report_b_address = edw2_url
        if side_by_side:
            comparison = Comparison(sources.PickerReport(picker_url=report_a_address,
                                                         report_name=edw3_report_name,
                                                         currency=side_by_side["currency"],
                                                         request_object=edw3_request_object),
                                    sources.PickerReport(picker_url=report_b_address,
                                                         report_name=edw2_report_name,
                                                         currency=side_by_side["currency"],
                                                         request_object=edw2_request_object)
                                    )
        else:
            comparison = Comparison(sources.PickerReport(picker_url=report_a_address,
                                                         report_name=edw3_report_name,
                                                         request_object=edw3_request_object),
                                    sources.PickerReport(picker_url=report_b_address,
                                                         report_name=edw2_report_name,
                                                         request_object=edw2_request_object)
                                    )
        comparison_start_date = start
        comparison_end_date = end
        try:
            comparison_start_date = modified_start
        except UnboundLocalError:
            pass
        try:
            comparison_end_date = modified_end
        except UnboundLocalError:
            pass
        if interval == 'day':
            comparison_end_date = comparison_start_date

        comparison.set_outputs(output_xlsx=output_xlsx, interval=interval, simple_report_name=report_name,
                               comparison_start_date=comparison_start_date, comparison_end_date=comparison_end_date,
                               cascade=cascade,
                               # side_by_side=side_by_side,
                               cascade_start_date=cascade_start_date, cascade_end_date=cascade_end_date)
        await comparison.run_and_barf()
        return comparison


async def loop_reports(dates_hash, edw2_request_object, edw3_request_object,
                       semaphore=3, interval="day",
                       output_xlsx=False, report_name='custom_comparison', side_by_side=None,
                       cascade=False,
                       cascade_start_date=None, cascade_end_date=None, picker_address=None):
    global sem
    sem = asyncio.Semaphore(semaphore)
    futures = []
    for date_idx, date in enumerate(dates_hash["dates"]):
        if date_idx == len(dates_hash["dates"]) - 1 and interval != 'day':
            break
        edw2_request_copy = copy.deepcopy(edw2_request_object)
        edw3_request_copy = copy.deepcopy(edw3_request_object)
        futures.append(run_comparison(dates_hash, date_idx, edw2_request_copy, edw3_request_copy, interval=interval,
                                      output_xlsx=output_xlsx, report_name=report_name,
                                      cascade=cascade,
                                      side_by_side=side_by_side,
                                      cascade_start_date=cascade_start_date, cascade_end_date=cascade_end_date,
                                      picker_address=picker_address))
    result = await asyncio.gather(*futures)
    return result


async def cascade_months(start, end, edw2_request_object, edw3_request_object, cascade_name, side_by_side=None):
    interval = 'month'
    split_start = start.split('/')
    split_end = end.split('/')
    start_date = datetime.datetime(int(split_start[2]), int(split_start[0]), int(split_start[1]))
    end_date = datetime.datetime(int(split_end[2]), int(split_end[0]), int(split_end[1]))
    months_hash = generate_dates(start_date, end_date, interval)
    edw2_request_copy = copy.deepcopy(edw2_request_object)
    edw3_request_copy = copy.deepcopy(edw3_request_object)
    # print('EDW2 - RO === \n ', edw2_request_copy, '\n')
    # print('EDW3 - RO === \n ', edw3_request_copy, '\n')
    result = await loop_reports(months_hash, edw2_request_copy, edw3_request_copy, semaphore=3, interval="month",
                                output_xlsx=True, report_name=cascade_name, side_by_side=side_by_side,
                                cascade=True, cascade_start_date=start, cascade_end_date=end)

    for month_comparison in result:
        months_comp.append(month_comparison.simple_difference_comparison)
        print("appending month sbs, " , len(months_comp))
        await cascade_weeks(
            month_comparison.comparison_start_date,
            month_comparison.comparison_end_date,
            edw2_request_object,
            edw3_request_object,
            cascade_name,
            side_by_side=side_by_side
        )
    months_summary_dict["comparisons"].extend(result)
    print("months summary list length -- ", len(months_summary_dict["comparisons"]))
    print("weeks summary list length -- ", len(weeks_summary_dict["comparisons"]))
    print("days summary list length -- ", len(days_summary_dict["comparisons"]))
    summaries = [days_summary_dict, weeks_summary_dict, months_summary_dict]

    await write_summary(summaries, cascade_name)
    if side_by_side:
        await write_side_by_side(cascade_name)


async def cascade_weeks(start, end, edw2_request_object, edw3_request_object, cascade_name,
                        side_by_side=None):
    interval = 'week'
    split_start = start.split('/')
    split_end = end.split('/')
    start_date = datetime.datetime(int(split_start[2]), int(split_start[0]), int(split_start[1]))
    end_date = datetime.datetime(int(split_end[2]), int(split_end[0]), int(split_end[1]))
    weeks_hash = generate_dates(start_date, end_date, interval)
    edw2_request_copy = copy.deepcopy(edw2_request_object)
    edw3_request_copy = copy.deepcopy(edw3_request_object)

    result = await loop_reports(weeks_hash, edw2_request_copy, edw3_request_copy, semaphore=3, interval="week",
                                output_xlsx=True, report_name=cascade_name, side_by_side=side_by_side,
                                cascade=True, cascade_start_date=start, cascade_end_date=end)

    for week_comparison in result:
        weeks_comp.append(week_comparison.side_by_side)
        print("appending week sbs, ", len(weeks_comp))
        await cascade_days(
            week_comparison.comparison_start_date,
            week_comparison.comparison_end_date,
            edw2_request_object,
            edw3_request_object,
            cascade_name,
            side_by_side=side_by_side
        )

    weeks_summary_dict["comparisons"].extend(result)


async def cascade_days(start, end, edw2_request_object, edw3_request_object, cascade_name,
                       side_by_side=None):
    interval = 'day'
    split_start = start.split('/')
    split_end = end.split('/')
    start_date = datetime.datetime(int(split_start[2]), int(split_start[0]), int(split_start[1]))
    end_date = datetime.datetime(int(split_end[2]), int(split_end[0]), int(split_end[1]))
    days_hash = generate_dates(start_date, end_date, interval)
    edw2_request_copy = copy.deepcopy(edw2_request_object)
    edw3_request_copy = copy.deepcopy(edw3_request_object)

    result = await loop_reports(days_hash, edw2_request_copy, edw3_request_copy, semaphore=3, interval="day",
                                output_xlsx=True, report_name=cascade_name, side_by_side=side_by_side,
                                cascade=True, cascade_start_date=start, cascade_end_date=end)

    days_summary_dict["comparisons"].extend(result)
    if side_by_side:
        for day_comparison in result:
            days_comp.append(day_comparison.side_by_side)
            print("appending days sbs, ", len(days_comp))


async def write_summary(list_of_comparison_dicts, cascade_name):
    xlsx_name = './validation_outputs/xlsx/cascade/' + cascade_name + '/' + cascade_name + '_summary' '.xlsx'
    wb = xlsxwriter.Workbook(xlsx_name)

    format_r = wb.add_format({'bg_color': '#FFC7CE',
                              'font_color': '#9C0006'})

    format_g = wb.add_format({'bg_color': '#C6EFCE',
                              'font_color': '#006100'})

    for sum_idx, summary_dict in enumerate(list_of_comparison_dicts):

        summary_name = summary_dict["interval"] + "_summary"
        summary = wb.add_worksheet(summary_name)
        inverted_summary = wb.add_worksheet(summary_name + '_inverted')

        header_row = 0
        col = 1

        header_col = 0
        row = 1

        for comp_idx, comparison in enumerate(list_of_comparison_dicts[sum_idx]["comparisons"]):
            if comparison is None or comparison.passing_records is None:
                # FIXME @le: instead of leaving passing_records to None when the load fails, it would be ideal to have it set to 0 or have some indication of an error
                continue
            output_string = {
                "passing_records": len(comparison.passing_records),
                "edw2_mismatches": len(comparison.edw2_mismatches),
                "edw3_mismatches": len(comparison.edw3_mismatches),
                "edw2_orphans": len(comparison.edw2_mismatches),
                "edw3_orphans": len(comparison.edw3_orphans)
            }

            summary.write(header_row, col + 2,
                          list_of_comparison_dicts[sum_idx]["comparisons"][comp_idx].comparison_start_date)
            inverted_summary.write(row + 2, header_col,
                                   list_of_comparison_dicts[sum_idx]["comparisons"][comp_idx].comparison_start_date)

            if len(comparison.edw3_orphans) or \
                    len(comparison.edw2_orphans) or \
                    len(comparison.edw2_mismatches) or \
                    len(comparison.edw3_mismatches):
                cell_format = format_r
            else:
                cell_format = format_g
            #
            summary.write(1, comp_idx + 3, str(output_string), cell_format)
            inverted_summary.write(comp_idx + 3, 1, str(output_string), cell_format)
            col += 1
            row += 1

        summary.write(1, 0, list_of_comparison_dicts[0]["comparisons"][0].simple_report_name)
        #
        summary.write(1, 1,
                      str(list_of_comparison_dicts[0]["comparisons"][0].reports[0].request_object))
        summary.write(1, 2,
                      str(list_of_comparison_dicts[0]["comparisons"][0].reports[1].request_object))

        inverted_summary.write(0, 1, str(list_of_comparison_dicts[0]["comparisons"][0].simple_report_name))
        # Write RO #1

        inverted_summary.write(1, 1,
                               str(list_of_comparison_dicts[0]["comparisons"][0].reports[0].request_object))
        # Write RO #2

        inverted_summary.write(2, 1,
                               str(list_of_comparison_dicts[0]["comparisons"][0].reports[1].request_object))
        print("summary has been added -- ", summary_name)
    wb.close()


def upload_summary(cascade_name):
    summary_directory = './validation_outputs/xlsx/cascade/' + cascade_name + '/' + cascade_name + '_summary' + '.xlsx'
    summary_sheet_names = [
        'month_summary',
        'week_summary',
        'day_summary'
    ]

    scope = ["https://spreadsheets.google.com/feeds", 'https://www.googleapis.com/auth/spreadsheets',
             "https://www.googleapis.com/auth/drive.file", "https://www.googleapis.com/auth/drive"]
    creds = ServiceAccountCredentials.from_json_keyfile_name("creds.json", scope)
    client = gspread.authorize(creds)
    sheet = client.open("AvantLink Regression Validation Summary")  # Open the spreadhseet
    ######

    gsheet_month = sheet.worksheet("month_summary")
    gsheet_week = sheet.worksheet("week_summary")
    gsheet_day = sheet.worksheet("day_summary")

    gsheet_month_inverted = sheet.worksheet("month_summary_inverted")
    gsheet_week_inverted = sheet.worksheet("week_summary_inverted")
    gsheet_day_inverted = sheet.worksheet("day_summary_inverted")
    base = ['Report Name', 'EDW2 Req Obj.', 'EDW3 Req Obj.', 'Run Date', '']
    start_date = datetime.datetime(2020, 1, 1)
    end_date = datetime.datetime(2021, 12, 31)
    months_dates = base.copy()
    months_dates.extend(generate_dates(start_date, end_date, "month")["dates"])
    weeks_dates = base.copy()
    weeks_dates.extend(generate_dates(start_date, end_date, "week")["dates"])
    days_dates = base.copy()
    days_dates.extend(generate_dates(start_date, end_date, "day")["dates"])

    # Write uninverted summaries to gSheets
    # gsheet_month.insert_row(months_dates)
    # gsheet_week.insert_row(weeks_dates)
    # gsheet_day.insert_row(days_dates)

    # write inverted summary to gsheet

    days_dates.reverse()
    weeks_dates.reverse()
    months_dates.reverse()
    data = load_workbook(summary_directory)
    # Iterate through the data of each local spreadsheet summary
    for sheet_name in summary_sheet_names:
        row_list = data[sheet_name].iter_rows(values_only=True)
        for row_idx, row in enumerate(row_list):
            if row_idx == 0:
                pass
            else:
                current_date = arrow.utcnow().format()
                current_gsheet = sheet.worksheet(sheet_name)
                current_row_local = list(row)
                current_row_local.insert(3, current_date)
                current_row_local.insert(4, '')
                new_gsheet_row = current_gsheet.insert_row(current_row_local)


async def write_side_by_side(cascade_name):
    print("Lengths within Side_by_side(),", len(months_comp), len(weeks_comp), len(days_comp))
    list_of_comparison_dicts = [
        {'interval': 'month', 'dataframe_set_list': months_comp},
        {'interval': 'day', 'dataframe_set_list': days_comp},
        {'interval': 'week', 'dataframe_set_list': weeks_comp}
    ]

    # months_comp list[{}]
    xlsx_name = './validation_outputs/xlsx/cascade/' + cascade_name + '/' + \
                cascade_name + '_difference_summary' '.xlsx'
    storage = {"month": [], "week": [], "day": []}
    for idx_1, interval_obj in enumerate(list_of_comparison_dicts):
        SBS_gathered = {
            "merge": [],
            # "mismatches_combined": [],
            # "edw2_raw": [],
            # "edw3_raw": [],
            # "orphans_merged": [],
            # "matches": [],
            # "edw2_mismatches": [],
            # "edw3_mismatches": [],
            # "edw2_orphans": [],
            # "edw3_orphans": []
        }
        for sbs_set in interval_obj["dataframe_set_list"]:
            try:
                # FIXME
                for key in sbs_set:
                    SBS_gathered[key].append(sbs_set[key])
            except KeyError:
                pass
        for list_key in SBS_gathered.keys():
            # print('DFLIST === , ', df_list)
            # print("length == ", len(SBS_gathered[list_key]), list_of_comparison_dicts[idx_1]["interval"])
            result = pd.concat(SBS_gathered[list_key], ignore_index=True)
            sheet_name = list_of_comparison_dicts[idx_1]["interval"] + "_" + list_key
            storage[list_of_comparison_dicts[idx_1]["interval"]].append({"dataframe": result, "sheet_name": sheet_name})
    with pd.ExcelWriter(xlsx_name) as writer:
        for interval in storage:
            for idx, sheet in enumerate(storage[interval]):
                storage[interval][idx]["dataframe"].to_excel(writer, sheet_name=storage[interval][idx]["sheet_name"])

def run_cascade(start_date, end_date, edw2_request_object, edw3_request_object,
                report_name="custom_comparison", side_by_side=None):
    loop = asyncio.new_event_loop()
    sem = None
    try:
        loop.run_until_complete(
            cascade_months(
                start_date,
                end_date,
                edw2_request_object,
                edw3_request_object,
                report_name,
                side_by_side=side_by_side
            ))
    except KeyboardInterrupt:
        pass
    finally:
        loop.close()

    pass
